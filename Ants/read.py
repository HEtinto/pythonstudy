import openpyxl
import numpy as np


def readdata():
    s1 = []
    count = 0
    s2 = []
    # 打开excel文件,获取工作簿对象
    wb = openpyxl.load_workbook('example.xlsx')
    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    print(ws)
    print(ws['A1'])  # 获取A列的第一个对象
    print(ws['A1'].value)
    for i in range(1, 2139):
        a = 'A' + str(i)
        b = 'B' + str(i)
        print(ws[a])
        print(ws[b])
        a1 = ws[a]
        print('Row {}, Column {} is {}'.format(a1.row, a1.column, a1.value))  # 打印这个单元格对象所在的行列的数值和内容
        print('Cell {} is {}\n'.format(a1.coordinate, a1.value))  # 获取单元格对象的所在列的行数和值
        s1.extend(str(a1.value).split(";", -1))
        count += 1
        # = str(a1.value).split(";", -1)
        # print(s1)
        b1 = ws[b]
        s2.extend(str(b1.value).split(";", -1))
        print('Row {}, Column {} is {}'.format(b1.row, b1.column, b1.value))  # 打印这个单元格对象所在的行列的数值和内容
        print('Cell {} is {}\n'.format(b1.coordinate, b1.value))  # 获取单元格对象的所在列的行数和值


    print(s1)
    print(s2)
    num1 = np.array([count, 2])
    for i in s1:
        # print("输出i:{}".format(i))
        s = []
        s.extend(i.split(",", -1))
        print(s)
        for j in range(0,2):
            num1 = np.append(num1, float(s[j]))

    for i in s2:
        # print("输出i:{}".format(i))
        s = []
        s.extend(i.split(",", -1))
        print(s)
        for j in range(0,2):
            num1 = np.append(num1, float(s[j]))

    num1 = num1.reshape(6415, 2)
    print(num1)
    return num1